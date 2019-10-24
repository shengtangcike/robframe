from openpyxl import load_workbook
##读取excel
wb = load_workbook(filename="财务流水-个人账户-运营流水.xlsx")
ws = wb['个人账户-运营流水1']
flowcode = []
#流水号集合
for i in range(2,ws.max_row+1):
    flowcode.append(ws['A%s'%i].value)
# print(flowcode)
print("平台导出数据%s"%len(flowcode))
#读取廊坊对账
f = open("10442_RM_20190702_182A","r",encoding="utf-8")
data = f.readlines()
print("廊坊推送数据%s"%len(data))
n = 1
for i in data:
    fl = i.split(",")
    if fl[0] in flowcode:
        #判断充值方式
        linfl = flowcode.index(fl[0]) + 2
        if float(ws['D%s'%linfl].value.replace(",","").replace(".","")) - float(fl[4]) < 10:
            # print("运营对账，【流水号】%s，廊坊：%s；后台：%s，"%(fl[0],float(ws['D%s'%linfl].value.replace(",","").replace(".",""))*100,float(fl[4])))
            pass
        else:
            print("数据不一致的流水号%s"%fl[0])
            print("运营对账，【流水号】%s，廊坊：%s；后台：%s，" % (fl[0],float(fl[4]),float(ws['D%s' % linfl].value.replace(",", "").replace(".", "")) * 100))
    else:
        print("导出的excel中不存在流水号%s"%fl[0])
    n += 1

df = []
for i in data:
    fl = i.split(",")
    df.append(fl[0])
for i in flowcode:
    if i in df:
        pass
    else:
        print("平台导出的excel多出的流水%s,在廊坊推送没有" %i)











# if "7320190702205528552" in flowcode:
#     print(flowcode)
#     print(flowcode.index("7320190702205528552")+2)
#     print(ws['D4'].value)
#     print(ws['F4'].value)

# print(ws.max_row)
# print(ws.max_column)
# for row in ws.rows:
#     for cell in row:
#         print(cell.value)



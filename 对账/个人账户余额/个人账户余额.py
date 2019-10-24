from openpyxl import load_workbook
##读取excel
wb = load_workbook(filename="财务流水-个人账户-recharge.xlsx")
ws = wb['个人账户-recharge1']
flowcode = []
#流水号集合
for i in range(2,ws.max_row+1):
    flowcode.append(ws['A%s'%i].value)
# print(flowcode)
print("平台导出数据%s"%len(flowcode))
#读取廊坊对账
f = open("10442_RF_20190702_180A","r",encoding="utf-8")
data = f.readlines()
print("廊坊推送数据%s"%len(data))
n = 1
for i in data:
    fl = i.split(",")
    if fl[0] in flowcode:
        #判断充值方式
        linfl = flowcode.index(fl[0]) + 2
        if ws['D%s'%linfl].value in fl[6] and ws['F%s'%linfl].value.replace(",","").replace(".","") == fl[3]:
            print("充值对账【编号】%s，【流水号】%s，【充值方式】廊坊：%s；后台：%s，【充值金额】廊坊：%s，资金：%s"%(n,fl[0],fl[6],ws['D%s'%linfl].value,fl[3],ws['F%s'%linfl].value.replace(",","").replace(".","")))
            # pass
        else:
            print("数据不一致的流水号%s"%fl[0])
            # print("充值对账：流水号%s，充值方式：廊坊%s；后台%s，充值金额：廊坊%s，资金%s" % (fl[0], fl[6], ws['D%s' % linfl].value, fl[3], ws['F%s' % linfl].value.replace(",", "").replace(".", "")))
    else:
        print("导出的excel中不存在流水号%s"%fl[0])
    n += 1

df = []
for i in data:
    fl = i.split(",")
    df.append(fl[0])
for i in flowcode:
    if i in df:
        pass
    else:
        print("平台导出的excel多出的流水%s,在廊坊推送没有" %i)











# if "7320190702205528552" in flowcode:
#     print(flowcode)
#     print(flowcode.index("7320190702205528552")+2)
#     print(ws['D4'].value)
#     print(ws['F4'].value)

# print(ws.max_row)
# print(ws.max_column)
# for row in ws.rows:
#     for cell in row:
#         print(cell.value)



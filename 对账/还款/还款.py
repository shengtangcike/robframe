from openpyxl import load_workbook
##读取excel
wb = load_workbook(filename="财务流水-个人账户-还款流水.xlsx")
ws = wb['个人账户-还款流水1']
flowcode = {}
#流水号集合
#流水号集合
for i in range(2,ws.max_row+1):
    if ws['H%s'%i].value != "借款服务费":
        if ws['A%s'%i].value not in flowcode.keys():
            flowcode[ws['A%s'%i].value] = 0
        flowcode[ws['A%s'%i].value] +=  float(ws['G%s'%i].value.replace(",",""))
    # flowcode.append(ws['A%s'%i].value)
# print(flowcode)
print("平台导出数据%s"%len(flowcode))


#读取廊坊对账
f = open("10442_RD_20190703_505A","r",encoding="utf-8")
data = f.readlines()

rd = {}
for i in data:
    fl = i.split(",")
    if fl[0] not in rd.keys():
        rd[fl[0]] = 0
    ky = float(fl[5]) + float(fl[6])
    rd[fl[0]] += ky
# print(rd)
print("廊坊推送数据%s"%len(rd))


for i in rd.keys():
    if i in flowcode.keys():
        if rd[i] - flowcode[i]*100 < 10:
            # print("还款【流水号】%s，廊坊：%s；后台：%s"%(i,rd[i],flowcode[i]*100))
            pass
        else:
            print("还款数据不一致的流水号%s，廊坊：%s；后台：%s" % (i, rd[i], flowcode[i]*100))
    else:
        print("导出的excel中不存在流水号%s"%i)




for i in flowcode.keys():
    if i in rd.keys():
        pass
    else:
        print("平台导出的excel多出的流水%s,在廊坊推送没有" %i)








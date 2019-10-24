from openpyxl import load_workbook
##读取excel
wb = load_workbook(filename="财务流水-个人账户-放款流水.xlsx")
ws = wb['个人账户-放款流水1']
flowcode = {}
#流水号集合
for i in range(2,ws.max_row+1):
    if ws['A%s'%i].value not in flowcode.keys():
        flowcode[ws['A%s'%i].value] = 0
    flowcode[ws['A%s'%i].value] +=  float(ws['G%s'%i].value.replace(",",""))
    # flowcode.append(ws['A%s'%i].value)
print(flowcode)
print("平台导出数据%s"%len(flowcode))
#读取廊坊对账

f = open("10442_RG_20190703_335A","r",encoding="utf-8")
data = f.readlines()
print("廊坊推送数据%s"%len(data))
n = 1
for i in data:
    fl = i.split(",")
    if fl[0] in flowcode.keys():
        if (flowcode[fl[0]]*100 - float(fl[3])) < 10:
            print("放款对账【编号】%s，【流水号】%s，【放款金额】廊坊：%s；后台：%s"%(n,fl[0],fl[3],flowcode[fl[0]]*100))
            # pass
        else:
            print("数据不一致的流水号%s"%fl[0])
            print("放款对账【编号】%s，【流水号】%s，【放款金额】廊坊：%s；后台：%s" % (n, fl[0], fl[3], flowcode[fl[0]] * 100))
    else:
        print("导出的excel中不存在流水号%s"%fl[0])
    n += 1


df = []
for i in data:
    fl = i.split(",")
    df.append(fl[0])
for i in flowcode.keys():
    if i in df:
        pass
    else:
        print("平台导出的excel多出的流水%s,在廊坊推送没有" %i)











# if "7320190702205528552" in flowcode:
#     print(flowcode)
#     print(flowcode.index("7320190702205528552")+2)
#     print(ws['D4'].value)
#     print(ws['F4'].value)

# print(ws.max_row)
# print(ws.max_column)
# for row in ws.rows:
#     for cell in row:
#         print(cell.value)



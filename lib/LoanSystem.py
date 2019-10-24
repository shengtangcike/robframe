#coding=utf-8
import requests
import json
import time,datetime
from openpyxl import load_workbook
from common.ReadAndWriteFiles import ReadAndWriteFiles
from lib.MysqlDB import MysqlDB

class LoanSystem(object):
    '''
    去投网借款系统基本功能
    '''
    def loansyslogin(self,username,password):
        '''
        :return: 借贷系统登录session
        '''
        print("借贷系统登录")
        self.loan = requests.Session()
        self.rwf = ReadAndWriteFiles()
        self.loanurlbase =  self.rwf.ini_read( self.rwf.pathhosturl, "qwt_loan_url")
        url_deb = self.loanurlbase + "/operate/user/login"
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {"account": username, "password": password, "system":"" }
        resl = self.loan.post(url_deb, data=data, headers=headers)
        print(resl.text)

    def loaninfodebt(self,name,phone,idcard,rate,month,money,jkname,type = None):
        '''
        :param name: 借款人姓名
        :param phone: 借款人手机号
        :param idcard: 借款人身份证号
        :param rate: 产品利率
        :param month: 产品期限
        :param money: 金额
        :type money: 默认是False，True代表推送智享
        :return:进件列表导入债权
        '''
       #生成债权
        wb = load_workbook(filename=self.rwf.pathloansysinto)
        ws = wb['Sheet1']
        ws['J2'] = name
        ws['C2'] = jkname
        ws['W2'] = phone
        ws['B2'] = idcard
        ws['F2'] = rate
        ws['G2'] = month
        ws['D2'] = money
        wb.save(self.rwf.pathloansysinto)
        wb.close()
        # 借款管理—进件列表—导入
        url_into = self.loanurlbase + "/operate/entrylist/upLoadDataDaoRu"
        # head_into = {"Content-Type":"multipart/form-data_testenv"}   #有header失败，没有header却成功了
        files = {'file': open(self.rwf.pathloansysinto, 'rb')}
        resl = self.loan.post(url_into, files=files)
        # 获取进件编号，并存储在resourse中
        debtcode = json.loads(resl.text)["list"]["list"][0]["debtcode"]
        print("进件编号为%s" % debtcode)
        if type == "True":
            self.rwf.update_resourse("resourse","${debtCode_zx}", debtcode)
        else:
            self.rwf.update_resourse("resourse","${debtCode}", debtcode)

    def debitcheck(self,debtCode,cifname,third):
        '''
        :param debtCode: 债权编号
        :param cifname: 垫付机构
        :param third: 编号
        :return: 在进件列表对债权进行编辑，在审核列表审核
        '''
        print("进件列表—编辑")
        print("债权编号%s"%debtCode)
        url_che = self.loanurlbase + "/operate/entrylist/entryList"
        data_che = json.dumps({"page": 1, "rows": 10, "offset": 0, "debtcode": debtCode, "borrowername": "", "loanmode": "","status": "", "beginDate": "", "endDate": ""})
        head_che = {"Content-Type": "application/json"}
        resl = self.loan.post(url_che, data=data_che, headers=head_che)
        resl_dic = json.loads(resl.text)
        a = resl_dic["rows"][0]["oid"]
        # oid编辑债权的主键信息
        url_tj = self.loanurlbase + "/operate/entrylist/updateSubmitEntryList"
        data_tj = json.dumps({"oid": a,"cusMask": "尽管借款人能够履行合同，但存在一些可能对偿还产生不利影响的因素，如这些因素继续下去，借款人的偿还能力受到影响。还款意愿正常，收入、稳定性等各方面状况正常",
                              "firstRepaymentSource": "借款企业主管业务收入偿还","loanFinanceInfo": "经营良好","loanRepaymentInfo": "企业经营状况良好，无异常","overdueReport": "良好",
                              "repaymentGuarantee": "信用资质良好，有车有房可抵押","twoRepaymentSource": "抵押房产车产变现"})

        head_tj = {"Content-Type": "application/json"}
        resl = self.loan.post(url_tj, data=data_tj, headers=head_tj)
        # print(resl.text)
        if '"errorCode":0' in resl.text:
            print("qtw借款端登编辑成功")
        else:
            raise Exception("qtw借款端登编辑失败")
        print("审核列表—审核")
        url_sh = self.loanurlbase + "/operate/entryAuditlist/updateAudit"
        data_sh = json.dumps({"oid": a, "description": "同意", "grtpaymentaccountname": cifname, "grtpaymentaccountno": third})
        head_sh = {"Content-Type": "application/json"}
        resl = self.loan.post(url_sh, data=data_sh, headers=head_sh)
        if '"errorCode":0' in resl.text:
            pass
        else:
            raise Exception("审核失败")

    def lendingbid(self,debtCode):
        '''
        :param debtCode: 债权编号
        :return: 满标后放款
        '''
        print("=======================第一步，查询放款列表的债权%s=========================" % debtCode)
        url_check = self.loanurlbase + "/operate/makelist/MakeList"
        header2 = {"Content-Type": "application/json"}
        data2 = {"page": 1, "rows": 10, "offset": 0, "loanNo": debtCode, "cifName": "", "returnMethod": "",
                 "loanStatus": "", "beginDate": "", "endDate": "", "pactissueNo": "", "cifPhone": ""}
        data2 = json.dumps(data2)
        resl = self.loan.post(url_check, data=data2, headers=header2)
        print(resl.text)
        oid = json.loads(resl.text)["rows"][0]["oid"]

        print("======================第二步，债权放款==========================")
        url_loan = self.loanurlbase + "/operate/makelist/loan"
        header3 = {"Content-Type": "application/x-www-form-urlencoded"}
        data3 = {'oid': oid}
        resl = self.loan.post(url_loan, data=data3, headers=header3)
        print(resl.text)
        print("======================第三步，判断债权状态为放款成功==========================")
        for i in range(10):
            time.sleep(20)
            url_check = self.loanurlbase + "/operate/makelist/MakeList"
            header2 = {"Content-Type": "application/json"}
            data2 = {"page": 1, "rows": 10, "offset": 0, "loanNo": debtCode, "cifName": "", "returnMethod": "",
                     "loanStatus": "", "beginDate": "", "endDate": "", "pactissueNo": "", "cifPhone": ""}
            data2 = json.dumps(data2)
            resl = self.loan.post(url_check, data=data2, headers=header2)
            # print(resl.text)
            rels = json.loads(resl.text)["rows"][0]["loanStatus"]
            print(rels)
            if "00" == rels:
                print("放款成功")
                break
        bidNocore = self.rwf.read_resourse("resourse","${bidNocore}")
        db = MysqlDB()
        script = "SELECT order_id FROM p2p_invest_order WHERE pactissue_no = '%s' LIMIT 1;" % bidNocore
        result = db.select_return_A_data(script)
        self.rwf.update_resourse("resourse", "${order_id}", str(result[0]))

    def overdueoneday(self,debtCode,term_num):
        '''
        :param debtCode: 债权编号
        :param term_num: 逾期期数
        :return: 逾期一天
        '''
        print("==========逾期1天申请垫付===========")
        #读取当天时间
        should_date = time.strftime('%Y-%m-%d', time.localtime())
        #时间更改为1天前
        should_ymd = should_date.split("-")
        should_date_1 = self.getday(y=int(should_ymd[0]), m=int(should_ymd[1]), d=int(should_ymd[2]), n=-1)
        print("还款日期向今天前推1天的时间%s"%should_date_1)
        #更改数据，使还款计划的第二期逾期
        db = MysqlDB()
        db.sql_update('UPDATE qtw_loan_db.p2p_borrower_return_plan SET should_date = "%s" WHERE debtcode = "%s" AND term_num = "%s"'%(should_date_1,debtCode,term_num))
        print("==================第一步，逾期申请==================")
        self.croundjob()
        # 根据债权信息查询【申请列表】oid
        oid = 0
        for i in range(10):
            oid = db.select_return_A_data('SELECT oid FROM qtw_loan_db.p2p_exceed_time_return WHERE debtcode = "%s"' % debtCode)
            if oid != None:
                oid = str(oid[0])
                print("在第%s分钟生产oid的值"%i)
                break
            time.sleep(10)


        # 借贷系统登录，申请垫付
        header = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {"oid": oid}
        url_1 = self.loanurlbase + "/operate/advances/applyReport"
        rslt = self.loan.post(url_1, data=data, headers=header)
        print(rslt.text)
        if 'success' in rslt.text:
            print("申请垫付成功")
        else:
            raise Exception("申请垫付失败")

        print("==================第二步，机构垫付==================")
        time.sleep(20)
        url_2 = self.loanurlbase +  "/operate/makelist/padPay"
        rslt = self.loan.post(url_2, data=data, headers=header)
        print(rslt.text)
        if '操作成功' in rslt.text:
            print("逾期垫付成功")
        else:
            raise Exception("逾期垫付失败")
        time.sleep(10)

    def overduemanyday(self,debtCode,term_num):
        print("==================借款人逾期8天，生成罚息数据==================")
        # 时间更改为8天前
        should_date = time.strftime('%Y-%m-%d', time.localtime())
        should_ymd = should_date.split("-")
        should_date_8 = self.getday(y=int(should_ymd[0]), m=int(should_ymd[1]), d=int(should_ymd[2]), n=-8)
        print("第二期还款日期向今天前推8天的时间%s" % should_date_8)
        # 更改数据，使还款计划的第二期逾期8天
        db = MysqlDB()
        db.sql_update('UPDATE qtw_loan_db.p2p_borrower_return_plan SET should_date = "%s" WHERE debtcode = "%s" AND term_num = "%s"' % (should_date_8, debtCode, term_num))
        db.sql_update("UPDATE qtw_loan_db.p2p_exceed_time_return SET expect_back_date = '%s',exceed_date = '%s' WHERE debtcode = '%s' AND term_num = %s;" % (should_date_8,should_date,debtCode, term_num))
        self.croundjob()
        punish_amt = 0
        for i in range(10):
            # 查询罚息的值
            punish_amt = db.select_return_A_data('SELECT punish_amt FROM qtw_loan_db.p2p_exceed_time_return WHERE debtcode = "%s" AND term_num = "2"' % debtCode)
            punish_amt = str(punish_amt[0])
            print("逾期罚息的金额为%s" % punish_amt)
            if int(punish_amt) > 0:
                print("逾期之后生成罚息，耗时%s分钟"%i)
                break
            time.sleep(10)
        if int(punish_amt) == 0:
                raise Exception("逾期之后没有生成罚息")

    def getday(self,y=2017,m=8,d=15,n=0):
        the_date = datetime.datetime(y,m,d)
        result_date = the_date + datetime.timedelta(days=n)
        d = result_date.strftime('%Y-%m-%d')
        return d
    def croundjob(self):
        '''
        :return: 执行借款系统的定时任务
        '''
        url_croundjob = self.rwf.ini_read( self.rwf.pathhosturl, "qwt_loan_cround_url")
        loan_cround = requests.Session()
        #未还款逾期处理
        url_unRepay = url_croundjob + "/activeTuning/unRepaymentBeoverdueHandleJob"
        # 罚息计算
        url_beoverdueMoney = url_croundjob + "/activeTuning/beoverdueMoneyCalculationJob"
        r1 = loan_cround.post(url_unRepay)
        r2 = loan_cround.post(url_beoverdueMoney)
        print(url_unRepay)
        print(url_beoverdueMoney)
        print(r1)
        print(r2)






if __name__=="__main__":
    loansys = LoanSystem()
    loansys.loansyslogin("admin","A123456")
    loansys.croundjob()
    # debtCode = "jj2019081200004"
    # db = MysqlDB()
    # oid = db.select_return_A_data('SELECT oid FROM qtw_loan_db.p2p_exceed_time_return WHERE debtcode = "%s"' % debtCode)
    # print(oid)
    # print(oid != None,111111111111111111111)
    # oid = str(oid[0])
    # print(oid)
    # # loansys.loaninfodebt("测试出借人十三","19800000013","511423198907010113","0.1","6","10000")
    #http://192.168.1.71:8090/qtw-loan-job/activeTuning/beoverdueMoneyCalculationJob
    #http://192.168.1.71:8090/qtw-loan-jobactiveTuning/beoverdueMoneyCalculationJob